"""
app.py — Servidor Flask para Gestor de Sesiones
Hernán Gabriel Stagni · MP 11919
Ejecutar: python app.py
"""

import threading
import webbrowser
import time
from pathlib import Path
from flask import Flask, jsonify, request, send_from_directory
from gestor import (
    cargar_config, guardar_config,
    escanear, mover_archivo,
    actualizar_excel,
    notificar_windows, notificar_email,
    crear_carpeta_paciente, editar_paciente
)

# ── Configuración Flask ────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
app = Flask(__name__, static_folder=str(BASE_DIR))

# Estado global del temporizador
timer_thread = None
timer_activo = False
timer_evento = threading.Event()  # permite interrumpir el sleep

# Log compartido entre hilo auto y frontend
_log_buffer = []
_log_lock = threading.Lock()
_tabla_desactualizada = False  # señal para que el frontend refresque


def _log_auto(msg, nivel="info"):
    """Agrega una entrada al buffer de log visible en el frontend."""
    global _log_buffer
    from datetime import datetime
    entrada = {
        "ts": datetime.now().strftime("%H:%M:%S"),
        "msg": msg,
        "nivel": nivel  # info | ok | warn
    }
    with _log_lock:
        _log_buffer.append(entrada)
        if len(_log_buffer) > 100:
            _log_buffer = _log_buffer[-100:]
    print(f"[Auto] {msg}")

# ── Servir la interfaz HTML ────────────────────────────────────────────────────

@app.route("/")
def index():
    response = send_from_directory(BASE_DIR, "gestor_sesiones.html")
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


# ── API: Configuración ─────────────────────────────────────────────────────────

@app.route("/api/config", methods=["GET"])
def get_config():
    cfg = cargar_config()
    # No enviar la contraseña al frontend
    cfg_segura = dict(cfg)
    cfg_segura["email_password"] = "●●●●●●●●" if cfg.get("email_password") else ""
    return jsonify(cfg_segura)


@app.route("/api/config", methods=["POST"])
def set_config():
    global timer_activo
    data = request.json
    cfg = cargar_config()

    campos = [
        "origen", "destino", "excel",
        "escaneo_automatico", "intervalo_minutos",
        "movimiento_automatico", "fuzzy_umbral",
        "notif_windows", "notif_email", "email_destino"
    ]
    for campo in campos:
        if campo in data:
            cfg[campo] = data[campo]

    # Solo actualizar password si no es el placeholder
    if "email_password" in data and data["email_password"] != "●●●●●●●●":
        cfg["email_password"] = data["email_password"]

    guardar_config(cfg)

    # Reiniciar temporizador si cambió la configuración de escaneo
    _gestionar_timer(cfg)

    return jsonify({"ok": True})


# ── API: Escaneo ───────────────────────────────────────────────────────────────

@app.route("/api/escanear", methods=["GET"])
def api_escanear():
    cfg = cargar_config()
    print(f"[DEBUG] Escaneando origen: {cfg['origen']}")
    resultado = escanear(cfg)
    print(f"[DEBUG] Resultado: error={resultado.get('error')}, archivos={len(resultado.get('archivos',[]))}")
    if resultado.get('archivos'):
        print(f"[DEBUG] Primer archivo: {resultado['archivos'][0]}")
    return jsonify(resultado)


# ── API: Mover archivos ────────────────────────────────────────────────────────

@app.route("/api/mover", methods=["POST"])
def api_mover():
    """
    Body: { "archivos": [ {"nombre": "...", "carpeta_destino": "..."}, ... ] }
    """
    data = request.json
    cfg = cargar_config()
    archivos = data.get("archivos", [])

    resultados = []
    movidos_ok = []
    errores_excel = []

    for item in archivos:
        nombre = item.get("nombre", "")
        carpeta = item.get("carpeta_destino", "Varios")

        resultado = mover_archivo(nombre, carpeta, cfg)
        resultados.append(resultado)

        if resultado["ok"]:
            movidos_ok.append(resultado)
            excel_ok = actualizar_excel(resultado, cfg)
            if not excel_ok:
                msg = f"No se registró en Excel: {nombre}"
                print(f"[ADVERTENCIA] {msg}")
                errores_excel.append(msg)

    # Notificaciones si hubo movimientos exitosos
    if movidos_ok:
        n = len(movidos_ok)
        resumen = "\n".join(f"· {r['nombre']} → {r['carpeta']}" for r in movidos_ok[:3])
        if n > 3:
            resumen += f"\n· ...y {n-3} más"

        if cfg.get("notif_windows"):
            notificar_windows(
                f"🎙️ {n} grabación(es) movida(s)",
                resumen
            )

        if cfg.get("notif_email"):
            notificar_email(movidos_ok, cfg)

    return jsonify({
        "ok": True,
        "total": len(archivos),
        "exitosos": len(movidos_ok),
        "resultados": resultados,
        "errores_excel": errores_excel
    })


# ── API: Pacientes ─────────────────────────────────────────────────────────────

@app.route("/api/pacientes", methods=["GET"])
def get_pacientes():
    cfg = cargar_config()
    destino_base = Path(cfg["destino"])

    pacientes = []
    cfg_pacientes = cfg.get("pacientes", {})

    if destino_base.exists():
        for carpeta in sorted(destino_base.iterdir()):
            if not carpeta.is_dir() or carpeta.name == "Varios":
                continue
            mp3s = list(carpeta.glob("*.mp3"))
            datos_cfg = cfg_pacientes.get(carpeta.name, {})
            pacientes.append({
                "nombre": carpeta.name,
                "sesiones": len(mp3s),
                "aliases": datos_cfg.get("aliases", []),
                "estado": datos_cfg.get("estado", "Actual"),
                "ultima_sesion": _ultima_sesion(mp3s)
            })

    return jsonify({"pacientes": pacientes})


def _ultima_sesion(mp3s):
    if not mp3s:
        return "—"
    try:
        mas_reciente = max(mp3s, key=lambda f: f.stat().st_mtime)
        ts = mas_reciente.stat().st_mtime
        from datetime import datetime
        return datetime.fromtimestamp(ts).strftime("%d/%m/%Y")
    except Exception:
        return "—"


@app.route("/api/pacientes", methods=["POST"])
def crear_paciente():
    data = request.json
    cfg = cargar_config()
    nombre = data.get("nombre", "").strip()
    aliases = data.get("aliases", "")
    estado = data.get("estado", "Actual")

    if not nombre:
        return jsonify({"ok": False, "error": "Nombre requerido"})

    try:
        ruta = crear_carpeta_paciente(nombre, aliases, estado, cfg)
        return jsonify({"ok": True, "ruta": ruta})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/pacientes", methods=["PUT"])
def actualizar_paciente():
    data = request.json
    cfg = cargar_config()
    nombre_original = data.get("nombre_original", "").strip()
    nombre_nuevo = data.get("nombre", "").strip()
    aliases = data.get("aliases", "")
    estado = data.get("estado", "Actual")

    if not nombre_original or not nombre_nuevo:
        return jsonify({"ok": False, "error": "Nombre requerido"})

    try:
        ruta = editar_paciente(nombre_original, nombre_nuevo, aliases, estado, cfg)
        return jsonify({"ok": True, "ruta": ruta})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


# ── API: Historial Excel ───────────────────────────────────────────────────────

@app.route("/api/historial", methods=["GET"])
def get_historial():
    cfg = cargar_config()
    excel_path = Path(cfg["excel"])

    if not excel_path.exists():
        return jsonify({"movimientos": []})

    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        ws = wb["Movimientos"]
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0] if rows else []
        movimientos = []
        for row in rows[1:]:
            if any(row):
                movimientos.append(dict(zip(headers, row)))
        wb.close()
        return jsonify({"movimientos": movimientos})
    except Exception as e:
        return jsonify({"movimientos": [], "error": str(e)})


# ── API: Email de prueba ───────────────────────────────────────────────────────

@app.route("/api/test-email", methods=["POST"])
def test_email():
    cfg = cargar_config()
    data = request.json
    if "email_password" in data and data["email_password"] != "●●●●●●●●":
        cfg["email_password"] = data["email_password"]
    if "email_destino" in data:
        cfg["email_destino"] = data["email_destino"]

    ok = notificar_email([{
        "ok": True,
        "nombre": "prueba_260409.mp3",
        "carpeta": "Paciente Test",
        "duracion": "0:01:00",
        "tamano_mb": 1.0
    }], cfg)
    return jsonify({"ok": ok})


# ── API: Estado del servidor ───────────────────────────────────────────────────

@app.route("/api/explorar", methods=["POST"])
def explorar_carpeta():
    """Abre el explorador de Windows para seleccionar una carpeta."""
    import tkinter as tk
    from tkinter import filedialog
    data = request.json
    ruta_actual = data.get("ruta_actual", "")
    campo = data.get("campo", "origen")

    try:
        root = tk.Tk()
        root.withdraw()
        root.wm_attributes('-topmost', True)
        carpeta = filedialog.askdirectory(
            title=f"Seleccionar carpeta {'origen' if campo=='origen' else 'destino'}",
            initialdir=ruta_actual if ruta_actual else "/"
        )
        root.destroy()
        if carpeta:
            # Convertir a formato Windows
            carpeta = carpeta.replace("/", "\\")
            cfg = cargar_config()
            cfg[campo] = carpeta
            guardar_config(cfg)
            return jsonify({"ok": True, "ruta": carpeta})
        return jsonify({"ok": False, "ruta": ""})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})



@app.route("/api/log/recientes", methods=["GET"])
def get_log_recientes():
    global _tabla_desactualizada
    desde = request.args.get("desde", 0, type=int)
    with _log_lock:
        nuevos = _log_buffer[desde:]
        total = len(_log_buffer)
    refresh = _tabla_desactualizada
    if _tabla_desactualizada:
        _tabla_desactualizada = False
    return jsonify({"entradas": nuevos, "total": total, "refresh": refresh})


@app.route("/api/abrir-excel", methods=["POST"])
def abrir_excel():
    cfg = cargar_config()
    excel_path = Path(cfg["excel"])
    if not excel_path.exists():
        return jsonify({"ok": False, "error": "El archivo Excel no existe aún. Mové al menos un archivo primero."})
    try:
        import os
        os.startfile(str(excel_path))
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/historial/stats", methods=["GET"])
def get_historial_stats():
    cfg = cargar_config()
    excel_path = Path(cfg["excel"])
    stats = {"total": 0, "este_mes": 0, "horas": "0h", "pacientes": 0}

    if not excel_path.exists():
        return jsonify(stats)

    try:
        import openpyxl
        from datetime import datetime
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        ws = wb["Movimientos"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) <= 1:
            return jsonify(stats)

        ahora = datetime.now()
        mes_actual = ahora.month
        anio_actual = ahora.year
        total = 0
        este_mes = 0
        pacientes_set = set()
        segundos_total = 0

        for row in rows[1:]:
            if not any(row):
                continue
            estado = str(row[7] or "") if len(row) > 7 else ""
            if not estado.startswith("✓"):
                continue
            total += 1

            # Fecha movimiento (col 0): "dd/mm/yyyy HH:MM"
            fecha_str = str(row[0] or "") if row[0] else ""
            try:
                fecha = datetime.strptime(fecha_str, "%d/%m/%Y %H:%M")
                if fecha.month == mes_actual and fecha.year == anio_actual:
                    este_mes += 1
            except Exception:
                pass

            # Paciente (col 2)
            pac = str(row[2] or "").strip() if len(row) > 2 else ""
            if pac and pac != "—":
                pacientes_set.add(pac)

            # Duración (col 4): "H:MM:SS"
            dur = str(row[4] or "") if len(row) > 4 else ""
            try:
                partes = dur.split(":")
                if len(partes) == 3:
                    segundos_total += int(partes[0]) * 3600 + int(partes[1]) * 60 + int(partes[2])
            except Exception:
                pass

        horas = segundos_total // 3600
        mins = (segundos_total % 3600) // 60
        horas_str = f"{horas}h {mins}m" if mins else f"{horas}h"

        stats = {
            "total": total,
            "este_mes": este_mes,
            "horas": horas_str,
            "pacientes": len(pacientes_set),
            "mes_nombre": ahora.strftime("%B %Y").lower()
        }
    except Exception as e:
        stats["error"] = str(e)

    return jsonify(stats)


@app.route("/api/estado", methods=["GET"])
def get_estado():
    cfg = cargar_config()
    return jsonify({
        "servidor": True,
        "escaneo_automatico": cfg.get("escaneo_automatico", False),
        "movimiento_automatico": cfg.get("movimiento_automatico", False),
        "intervalo_minutos": cfg.get("intervalo_minutos", 30)
    })


# ── Temporizador de escaneo automático ────────────────────────────────────────

def _loop_escaneo():
    global timer_activo, _tabla_desactualizada
    _log_auto("Escaneo automático activado.", "info")
    while timer_activo:
        cfg = cargar_config()
        if not cfg.get("escaneo_automatico"):
            _log_auto("Escaneo automático desactivado.", "warn")
            break

        intervalo_min = cfg.get("intervalo_minutos", 30)
        _log_auto(f"[Auto] Escaneando carpeta origen...", "info")

        resultado = escanear(cfg)
        archivos = resultado.get("archivos", [])
        listos = [a for a in archivos if a["tipo"] == "ok"]
        sin_renombrar = [a for a in archivos if a["tipo"] == "sin_renombrar"]
        varios = [a for a in archivos if a["tipo"] == "varios"]

        if cfg.get("movimiento_automatico"):
            if listos:
                movidos_ok = []
                for a in listos:
                    r = mover_archivo(a["nombre"], a["carpeta_destino"], cfg)
                    if r["ok"]:
                        excel_ok = actualizar_excel(r, cfg)
                        movidos_ok.append(r)
                        if excel_ok:
                            _log_auto(f"[Auto] ✓ {r['nombre']} → {r['carpeta']} (Excel OK)", "ok")
                        else:
                            _log_auto(f"[Auto] ✓ {r['nombre']} → {r['carpeta']} (⚠ NO se registró en Excel — ¿está abierto?)", "warn")
                    else:
                        _log_auto(f"[Auto] ✗ {r.get('error','Error')}", "warn")

                _tabla_desactualizada = True

                if movidos_ok and cfg.get("notif_windows"):
                    notificar_windows(
                        f"🎙️ {len(movidos_ok)} grabación(es) movida(s) automáticamente",
                        "\n".join(f"· {r['nombre']} → {r['carpeta']}" for r in movidos_ok[:3])
                    )
                if movidos_ok and cfg.get("notif_email"):
                    notificar_email(movidos_ok, cfg)
            else:
                _log_auto(
                    f"[Auto] {len(listos)} listos · {len(sin_renombrar)} sin renombrar · {len(varios)} sin clasificar",
                    "info"
                )
        else:
            _log_auto(
                f"[Auto] {len(listos)} listo(s) para mover · {len(sin_renombrar)} sin renombrar · {len(varios)} sin clasificar",
                "info"
            )
            _tabla_desactualizada = True
            if listos and cfg.get("notif_windows"):
                notificar_windows(
                    f"🎙️ {len(listos)} grabación(es) lista(s) para mover",
                    "Abrí el Gestor de Sesiones para confirmar el movimiento."
                )

        _log_auto(f"[Auto] Próximo escaneo en {intervalo_min} min.", "info")
        timer_evento.clear()
        timer_evento.wait(timeout=intervalo_min * 60)

    _log_auto("Hilo de escaneo detenido.", "info")


def _gestionar_timer(cfg):
    global timer_thread, timer_activo
    # Siempre detener el hilo existente para aplicar la nueva config
    timer_activo = False
    timer_evento.set()  # despierta el sleep si está esperando
    if timer_thread and timer_thread.is_alive():
        timer_thread.join(timeout=3)

    if cfg.get("escaneo_automatico"):
        timer_activo = True
        timer_evento.clear()
        timer_thread = threading.Thread(target=_loop_escaneo, daemon=True)
        timer_thread.start()
        print(f"[Timer] Timer reiniciado. Intervalo: {cfg.get('intervalo_minutos',30)} min.")


# ── Inicio ─────────────────────────────────────────────────────────────────────

def abrir_navegador():
    time.sleep(1.2)
    webbrowser.open("http://localhost:5000")


if __name__ == "__main__":
    cfg_inicial = cargar_config()
    _gestionar_timer(cfg_inicial)

    # Abrir navegador automáticamente
    threading.Thread(target=abrir_navegador, daemon=True).start()

    print("=" * 50)
    print("  Gestor de Sesiones — Hernán Gabriel Stagni")
    print("  Servidor corriendo en http://localhost:5000")
    print("  Cerrá esta ventana para detener el sistema.")
    print("=" * 50)

    app.run(host="127.0.0.1", port=5000, debug=False, use_reloader=False)
