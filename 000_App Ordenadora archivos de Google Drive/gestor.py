"""
gestor.py — Motor de escaneo y movimiento de grabaciones
Hernán Gabriel Stagni · MP 11919
"""

import os
import re
import shutil
import hashlib
import json
import smtplib
import unicodedata
from pathlib import Path
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

try:
    from mutagen.mp3 import MP3
    MUTAGEN_OK = True
except ImportError:
    MUTAGEN_OK = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
    print("[ADVERTENCIA] openpyxl no está instalado. El registro Excel está DESACTIVADO.")
    print("  → Ejecutá: pip install openpyxl")

try:
    from win10toast import ToastNotifier
    TOAST_OK = True
except ImportError:
    TOAST_OK = False

# ── Configuración por defecto ──────────────────────────────────────────────────
CONFIG_PATH = Path(__file__).parent / "config.json"

DEFAULT_CONFIG = {
    "origen": r"G:\Mi unidad\00_Easy Voice Recorder",
    "destino": r"B:\04_Grabaciones sesiones terapia\000_Sesiones de terapia PACIENTES\01_Actuales",
    "excel": r"B:\04_Grabaciones sesiones terapia\000_Sesiones de terapia PACIENTES\01_Actuales\000_Registro sesiones.xlsx",
    "escaneo_automatico": False,
    "intervalo_minutos": 30,
    "movimiento_automatico": False,
    "fuzzy_umbral": 85,
    "notif_windows": True,
    "notif_email": False,
    "email_destino": "psicologohernanstagni@gmail.com",
    "email_password": "",
    "pacientes": {}
}

# ── Helpers ────────────────────────────────────────────────────────────────────

def cargar_config():
    if CONFIG_PATH.exists():
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            # Rellenar claves faltantes con defaults
            for k, v in DEFAULT_CONFIG.items():
                cfg.setdefault(k, v)
            return cfg
        except Exception:
            pass
    return dict(DEFAULT_CONFIG)


def guardar_config(cfg):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def normalizar(texto):
    """Minúsculas, sin acentos, sin espacios dobles."""
    texto = texto.strip().lower()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    texto = re.sub(r"\s+", " ", texto)
    return texto


def similitud(a, b):
    """Porcentaje de similitud simple basada en caracteres comunes (Dice coefficient)."""
    a, b = normalizar(a), normalizar(b)
    if a == b:
        return 100
    if not a or not b:
        return 0
    set_a = set(a[i:i+2] for i in range(len(a)-1))
    set_b = set(b[i:i+2] for i in range(len(b)-1))
    if not set_a or not set_b:
        return 0
    interseccion = len(set_a & set_b)
    return int(200 * interseccion / (len(set_a) + len(set_b)))


# Patrón: 6 dígitos (AAMMDD) + espacio + nombre
PATRON_VALIDO = re.compile(r"^(\d{6})\s+(.+?)\.mp3$", re.IGNORECASE)


def clasificar_archivo(nombre):
    """
    Devuelve dict con tipo y datos del archivo.
    tipo: 'ok' | 'varios' | 'sin_renombrar'
    """
    if re.match(r"^mi grabaci[oó]n\s*\d+\.mp3$", nombre, re.IGNORECASE):
        return {"tipo": "sin_renombrar", "paciente": "", "fecha_str": ""}

    m = PATRON_VALIDO.match(nombre)
    if m:
        fecha_str = m.group(1)
        paciente_raw = m.group(2).strip()
        return {"tipo": "ok", "paciente": paciente_raw, "fecha_str": fecha_str}

    return {"tipo": "varios", "paciente": "", "fecha_str": ""}


def obtener_duracion(ruta):
    """Duración del MP3 en formato H:MM:SS."""
    if not MUTAGEN_OK:
        return "—"
    try:
        audio = MP3(str(ruta))
        seg = int(audio.info.length)
        h, rem = divmod(seg, 3600)
        m, s = divmod(rem, 60)
        return f"{h}:{m:02d}:{s:02d}" if h else f"0:{m:02d}:{s:02d}"
    except Exception:
        return "—"


def obtener_tamano_mb(ruta):
    try:
        return round(ruta.stat().st_size / (1024 * 1024), 1)
    except Exception:
        return 0


def checksum(ruta):
    h = hashlib.md5()
    with open(ruta, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def fecha_desde_str(fecha_str):
    """Convierte AAMMDD → datetime."""
    try:
        return datetime.strptime(fecha_str, "%y%m%d")
    except Exception:
        return None


# ── Escaneo ───────────────────────────────────────────────────────────────────

def escanear(cfg):
    """
    Escanea la carpeta origen y devuelve lista de archivos clasificados.
    Cada item: {nombre, tipo, paciente, carpeta_destino, duracion, tamano_mb,
                fecha_grabacion, ruta_completa}
    """
    origen = Path(cfg["origen"])
    destino_base = Path(cfg["destino"])
    umbral = cfg.get("fuzzy_umbral", 85)
    pacientes_cfg = cfg.get("pacientes", {})

    if not origen.exists():
        return {"error": f"Carpeta origen no encontrada: {origen}", "archivos": []}

    carpetas_existentes = []
    if destino_base.exists():
        carpetas_existentes = [d.name for d in destino_base.iterdir() if d.is_dir()]

    archivos = []
    for f in origen.iterdir():
        if not f.is_file() or f.suffix.lower() != ".mp3":
            continue

        info = clasificar_archivo(f.name)
        duracion = obtener_duracion(f)
        tamano = obtener_tamano_mb(f)
        fecha_grab = None
        if info["fecha_str"]:
            fecha_grab = fecha_desde_str(info["fecha_str"])

        carpeta_dest = "Varios"

        if info["tipo"] == "ok":
            paciente_raw = info["paciente"]
            # Buscar en aliases del config
            nombre_oficial = None
            for nombre_carp, datos in pacientes_cfg.items():
                aliases = [normalizar(a) for a in datos.get("aliases", [])]
                aliases.append(normalizar(nombre_carp))
                if normalizar(paciente_raw) in aliases:
                    nombre_oficial = nombre_carp
                    break

            # Si no encontró en aliases, buscar por fuzzy en carpetas existentes
            if not nombre_oficial:
                mejor = None
                mejor_score = 0
                for carp in carpetas_existentes:
                    if carp == "Varios":
                        continue
                    score = similitud(paciente_raw, carp)
                    if score > mejor_score:
                        mejor_score = score
                        mejor = carp
                if mejor and mejor_score >= umbral:
                    nombre_oficial = mejor

            # Si aún no encontró, usar el nombre raw (se creará la carpeta)
            if not nombre_oficial:
                nombre_oficial = paciente_raw

            carpeta_dest = nombre_oficial

        archivos.append({
            "nombre": f.name,
            "tipo": info["tipo"],
            "paciente": info["paciente"] if info["tipo"] == "ok" else "",
            "paciente_detectado": carpeta_dest if info["tipo"] == "ok" else "",
            "carpeta_destino": carpeta_dest,
            "duracion": duracion,
            "tamano_mb": tamano,
            "fecha_grabacion": fecha_grab.strftime("%d/%m/%Y") if fecha_grab else "—",
            "ruta_completa": str(f)
        })

    orden = {'ok': 0, 'sin_renombrar': 1, 'varios': 2}
    archivos.sort(key=lambda x: orden.get(x['tipo'], 1))

    return {"error": None, "archivos": archivos}


# ── Movimiento seguro ──────────────────────────────────────────────────────────

def mover_archivo(nombre_archivo, carpeta_destino_nombre, cfg, log_callback=None):
    """
    Mueve un archivo de origen a destino de forma segura:
    1. Copia al destino
    2. Verifica checksum
    3. Elimina el original
    Devuelve dict con resultado.
    """
    origen = Path(cfg["origen"]) / nombre_archivo
    destino_base = Path(cfg["destino"])
    carpeta_final = destino_base / carpeta_destino_nombre

    def log(msg):
        if log_callback:
            log_callback(msg)

    if not origen.exists():
        return {"ok": False, "error": f"Archivo no encontrado: {nombre_archivo}"}

    try:
        # Crear carpeta si no existe
        carpeta_final.mkdir(parents=True, exist_ok=True)

        destino_archivo = carpeta_final / nombre_archivo

        # Si ya existe en destino, no duplicar
        if destino_archivo.exists():
            return {"ok": False, "error": f"Ya existe en destino: {nombre_archivo}"}

        # Calcular checksum del original
        cs_original = checksum(origen)

        # Copiar
        shutil.copy2(str(origen), str(destino_archivo))

        # Verificar integridad
        cs_copia = checksum(destino_archivo)
        if cs_original != cs_copia:
            destino_archivo.unlink()
            return {"ok": False, "error": f"Error de integridad en la copia: {nombre_archivo}"}

        # Eliminar original
        origen.unlink()

        tamano = obtener_tamano_mb(destino_archivo)
        duracion = obtener_duracion(destino_archivo)

        # Extraer paciente y fecha de grabación del nombre del archivo
        info = clasificar_archivo(nombre_archivo)
        paciente = info.get("paciente", "") or carpeta_destino_nombre
        fecha_grab = ""
        if info.get("fecha_str"):
            fg = fecha_desde_str(info["fecha_str"])
            if fg:
                fecha_grab = fg.strftime("%d/%m/%Y")

        log(f"✓ {nombre_archivo} → {carpeta_destino_nombre} ({duracion} · {tamano} MB)")

        return {
            "ok": True,
            "nombre": nombre_archivo,
            "carpeta": carpeta_destino_nombre,
            "paciente": paciente,
            "fecha_grabacion": fecha_grab,
            "ruta_destino": str(destino_archivo),
            "tamano_mb": tamano,
            "duracion": duracion
        }

    except Exception as e:
        return {"ok": False, "error": str(e)}


# ── Registro Excel ─────────────────────────────────────────────────────────────

VERDE_OSCURO = "1E3D2B"
DORADO = "C8922A"
CREMA = "FAF7F2"
NARANJA_CLARO = "FFF6E4"

PENDING_CSV = Path(__file__).parent / "_pendientes_excel.csv"

HEADERS = [
    "Fecha movimiento", "Nombre archivo", "Paciente",
    "Carpeta destino", "Duración", "Tamaño (MB)",
    "Fecha grabación", "Estado"
]


def _fila_desde_resultado(resultado):
    estado = "✓ OK" if resultado["ok"] else f"✗ {resultado.get('error','Error')}"
    return [
        datetime.now().strftime("%d/%m/%Y %H:%M"),
        resultado.get("nombre", "—"),
        resultado.get("paciente", resultado.get("carpeta", "—")),
        resultado.get("carpeta", "—"),
        resultado.get("duracion", "—"),
        resultado.get("tamano_mb", 0),
        resultado.get("fecha_grabacion", "—"),
        estado
    ]


def _guardar_pendiente_csv(fila):
    """Guarda una fila en el CSV de pendientes cuando Excel está bloqueado."""
    import csv
    escribir_header = not PENDING_CSV.exists()
    with open(PENDING_CSV, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        if escribir_header:
            w.writerow(HEADERS)
        w.writerow(fila)
    print(f"[Excel] Registro pendiente guardado en CSV: {fila[1]}")


def _incorporar_pendientes(wb):
    """Si existe el CSV de pendientes, lo incorpora al Excel y lo elimina."""
    import csv
    if not PENDING_CSV.exists():
        return
    try:
        ws = wb["Movimientos"]
        with open(PENDING_CSV, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)
        # rows[0] es el header, saltar
        pendientes = rows[1:] if len(rows) > 1 else []
        for fila in pendientes:
            if any(fila):
                ws.append(fila)
                _aplicar_estilo_fila(ws, ws.max_row, ok=str(fila[7]).startswith("✓"))
        if pendientes:
            print(f"[Excel] {len(pendientes)} registro(s) pendiente(s) incorporados desde CSV.")
        PENDING_CSV.unlink()
    except Exception as e:
        print(f"[Excel] No se pudieron incorporar pendientes: {e}")


def _aplicar_estilo_fila(ws, fila_num, ok=True):
    fill = PatternFill("solid", fgColor=CREMA if ok else "FDF0F0")
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(fila_num, col)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
        cell.border = Border(bottom=Side(style="thin", color="D8D0C0"))


def _abrir_o_crear_excel(excel_path):
    """Abre el Excel existente o crea uno nuevo con las hojas base."""
    if excel_path.exists():
        wb = openpyxl.load_workbook(excel_path)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Movimientos"
        _crear_encabezado(ws, HEADERS)
        ws2 = wb.create_sheet("Pacientes")
        _crear_encabezado(ws2, ["Nombre carpeta (oficial)", "Variantes / aliases", "Estado"])
        ws3 = wb.create_sheet("Resumen")
        _crear_encabezado(ws3, ["Paciente", "Total sesiones", "Duración acumulada", "Última sesión"])
        print(f"[Excel] Archivo creado en: {excel_path}")

    # Asegurar que exista la hoja Movimientos
    if "Movimientos" not in wb.sheetnames:
        ws = wb.create_sheet("Movimientos", 0)
        _crear_encabezado(ws, HEADERS)

    return wb


def actualizar_excel(resultado, cfg):
    """Agrega una fila al Excel. Si está bloqueado, guarda en CSV y reintenta la próxima vez."""
    if not OPENPYXL_OK:
        print("[ERROR] openpyxl no instalado. Ejecutá: pip install openpyxl")
        return False

    excel_path = Path(cfg["excel"])
    try:
        excel_path.parent.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        print(f"[ERROR Excel] No se pudo crear la carpeta: {e}")
        return False

    fila = _fila_desde_resultado(resultado)

    # Intentar abrir y guardar (hasta 3 reintentos con 1s de pausa)
    for intento in range(3):
        try:
            wb = _abrir_o_crear_excel(excel_path)
            ws = wb["Movimientos"]

            # Incorporar registros pendientes del CSV antes de agregar el nuevo
            _incorporar_pendientes(wb)

            ws.append(fila)
            _aplicar_estilo_fila(ws, ws.max_row, ok=resultado["ok"])

            anchos = [18, 35, 20, 50, 10, 12, 15, 20]
            for i, ancho in enumerate(anchos, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

            wb.save(excel_path)
            print(f"[Excel] ✓ Guardado: {resultado.get('nombre','?')} → {resultado.get('carpeta','?')}")
            return True

        except PermissionError:
            if intento < 2:
                import time as _time
                print(f"[Excel] Archivo bloqueado, reintentando en 1s (intento {intento+1}/3)...")
                _time.sleep(1)
            else:
                # Después de 3 intentos: guardar en CSV pendiente
                _guardar_pendiente_csv(fila)
                print("[Excel] Excel sigue bloqueado. Registro guardado en CSV pendiente.")
                print("  → Se incorporará automáticamente la próxima vez que el Excel esté libre.")
                return False

        except Exception as e:
            print(f"[ERROR Excel] {e}")
            _guardar_pendiente_csv(fila)
            return False

    return False


def _crear_encabezado(ws, headers):
    ws.append(headers)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=VERDE_OSCURO)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="medium", color=DORADO))
    ws.row_dimensions[1].height = 22


# ── Notificaciones ─────────────────────────────────────────────────────────────

def notificar_windows(titulo, mensaje):
    if not TOAST_OK:
        return
    try:
        toaster = ToastNotifier()
        toaster.show_toast(titulo, mensaje, duration=8, threaded=True)
    except Exception:
        pass


def notificar_email(movidos, cfg):
    """Envía email resumen de archivos movidos."""
    if not cfg.get("notif_email") or not cfg.get("email_password"):
        return False
    try:
        destinatario = cfg["email_destino"]
        password = cfg["email_password"]
        fecha = datetime.now().strftime("%d/%m/%Y")
        asunto = f"{len(movidos)} grabación(es) movida(s) — {fecha}"

        lineas = "\n".join(
            f"· {r['nombre']} → {r['carpeta']} ({r['duracion']} · {r['tamano_mb']} MB)"
            for r in movidos if r.get("ok")
        )

        cuerpo = f"""Se procesaron automáticamente los siguientes archivos:

{lineas}

---
Gestor de Sesiones · Hernán Gabriel Stagni · MP 11919
"""
        msg = MIMEMultipart()
        msg["From"] = destinatario
        msg["To"] = destinatario
        msg["Subject"] = asunto
        msg.attach(MIMEText(cuerpo, "plain", "utf-8"))

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(destinatario, password)
            server.sendmail(destinatario, destinatario, msg.as_string())
        return True
    except Exception as e:
        print(f"Error email: {e}")
        return False


# ── Crear carpeta paciente ─────────────────────────────────────────────────────

def crear_carpeta_paciente(nombre, aliases, estado, cfg):
    """Crea carpeta física y registra en config."""
    destino_base = Path(cfg["destino"])
    if estado == "Inactivo":
        # Un nivel arriba → 02_Inactivos
        carpeta_inactivos = destino_base.parent / "02_Inactivos"
        carpeta_inactivos.mkdir(parents=True, exist_ok=True)
        nueva = carpeta_inactivos / nombre
    else:
        nueva = destino_base / nombre

    nueva.mkdir(parents=True, exist_ok=True)

    cfg.setdefault("pacientes", {})[nombre] = {
        "aliases": [a.strip() for a in aliases.split(",") if a.strip()],
        "estado": estado
    }
    guardar_config(cfg)
    return str(nueva)


def editar_paciente(nombre_original, nombre_nuevo, aliases, estado, cfg):
    """Renombra carpeta, actualiza aliases/estado y registra en config."""
    import shutil
    destino_base = Path(cfg["destino"])
    carpeta_inactivos = destino_base.parent / "02_Inactivos"

    # Buscar la carpeta actual del paciente
    ruta_en_act   = destino_base / nombre_original
    ruta_en_inact = carpeta_inactivos / nombre_original

    if ruta_en_act.exists():
        ruta_actual = ruta_en_act
    elif ruta_en_inact.exists():
        ruta_actual = ruta_en_inact
    else:
        raise FileNotFoundError(f"No se encontró la carpeta del paciente: {nombre_original}")

    # Carpeta destino según el nuevo estado
    if estado == "Inactivo":
        carpeta_inactivos.mkdir(parents=True, exist_ok=True)
        ruta_nueva = carpeta_inactivos / nombre_nuevo
    else:
        ruta_nueva = destino_base / nombre_nuevo

    # Mover/renombrar si cambia la ruta
    if ruta_actual != ruta_nueva:
        if ruta_nueva.exists():
            raise FileExistsError(f"Ya existe una carpeta con el nombre: {nombre_nuevo}")
        shutil.move(str(ruta_actual), str(ruta_nueva))

    # Actualizar config
    pacientes = cfg.setdefault("pacientes", {})
    if nombre_original in pacientes:
        del pacientes[nombre_original]
    pacientes[nombre_nuevo] = {
        "aliases": [a.strip() for a in aliases.split(",") if a.strip()],
        "estado": estado
    }
    guardar_config(cfg)
    return str(ruta_nueva)
